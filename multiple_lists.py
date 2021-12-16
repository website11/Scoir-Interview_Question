import openpyxl
import csv

# A list might be better to demonstrate sort/search algorithms



def swap(data, first, second):
    temp = data[first]
    data[first] = data[second]
    data[second] = temp


def partition(data, start, stop):
    pivot = start
    for i in range(start, stop):
        if data[i] > data[stop]:
            swap(data, pivot, i)
            pivot += 1
    swap(data, pivot, stop)

    return pivot


def quickSort(data, start, stop):
    if start < stop:
        p = partition(data, start, stop)
        quickSort(data, start, p - 1)
        quickSort(data, p + 1, stop)




def binarySearch(target, names, end, bias):
    start = 0
    stop = end - 1
    last = -1
    while start <= stop:
        middle = (start + stop) / 2
        if (target == names[middle]) & bias == 1:
            last = middle
            start = middle + 1

        elif (target == names[middle]) & bias == 0:
            last = middle
            start = middle - 1

        elif target < 0:
            stop = middle - 1

        else:
            start = middle + 1

    return last



def new_dictionary(data):
    rows = data.max_row
    names = []
    for i in range(1, rows + 1):
        first_name = data.cell(row=i, column=1).value
        last_name = data.cell(row=i, column=2).value
        dob = data.cell(row=i, column=3).value
        names.append([first_name, last_name,dob])

    return names


if __name__ == "__main__":
    file_path = input("Enter file path location: ")
    user_input = input("Search names by: (1) First Name | (2) Last Name | (3) Date of Birth\n Pick an option: ")

    csv_wb = openpyxl.load_workbook(file_path)
    table_data = csv_wb.active
    rows = table_data.max_row

    if user_input == "1":
        user_dob = int(input("Enter a date of birth (in YYYYMMDD): "))
        all_names = new_dictionary(table_data)
        quickSort(all_names, 0, rows - 1)
        if user_dob in all_names:
            result = all_names[user_dob]
            print(result)
        else:
            print("Sorry, this date of birth does not exist in the system!")
